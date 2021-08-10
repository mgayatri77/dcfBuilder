from openpyxl.utils.cell import get_column_letter
import yfinance as yf
from openpyxl.utils import get_column_letter
import urllib.request
import json
import constants
import time

# divides a percentage by 100. For example, 98% -> 0.98
# @param perc - the percentage
# @return float - number divided by 100.0
def percent_to_dec(perc):
    return perc/100.0

# converts column and row to excel format. For example 1,1 -> A1
# @param col - excel column number (1-based)
# @param row - excel row number (1-based)
# @ return string - cell number in excel 
def get_excel_cell(col, row): 
    return get_column_letter(col) + str(row)

# pulls stock price from ticker using yfinance
# @param ticker - stock ticker
# @return float - regular market stock price
def get_stock_price_from_ticker(ticker): 
    # remove whitespace and convert to uppercase
    ticker = ticker.strip().upper()
    stock_info = yf.Ticker(ticker).info

    #return stock price
    price_key = "regularMarketPrice" 
    return stock_info[price_key]

# gets JSON from given URL using urllib, catches HTTPErrors 
# if they occur during execution of request 
# since the SEC API fails often, this attempts to get data 
# from SEC 'MAX_ATTEMPTS' times before stopping
# @param url_link - URL string
# @return json - json_data as dictionary 
def get_json_from_url_link(url_link):
    # fetch and return JSON from url_link
    MAX_ATTEMPTS = 30
    num_tries = 1

    # try 'MAX_ATTEMPTS' times to fetch data
    while(num_tries < MAX_ATTEMPTS): 
        try: 
            with urllib.request.urlopen(url_link) as url:
                json_data = json.loads(url.read().decode())
                return json_data
        # catch HTTP Errors and print number of attempts completed
        except urllib.error.HTTPError as e: 
            print("Error fetching SEC JSON Data ->", e)
            print("Attempt " + str(num_tries) + " out of " + str(MAX_ATTEMPTS))
            num_tries += 1
            time.sleep(0.5)
    
    #return error code if URL failed to load        
    return constants.ERRORCODE

# gets CIK code from given stock ticker using SEC Ticker to CIK map
# @param ticker - stock ticker
# @return string - CIK code
def get_cik_from_ticker(ticker):
    # remove whitespace and convert to uppercase, set initial CIK to -1
    ticker = ticker.strip().upper()
    CIK = constants.ERRORCODE

    # lookup ticker from SEC Edgar API
    ticker_cik_map = get_json_from_url_link(constants.CIK_link)
    assert(ticker_cik_map != constants.ERRORCODE), "Unable to load Company Ticker to CIK JSON"
    for i in range(len(ticker_cik_map)):
        if(ticker == ticker_cik_map[str(i)]["ticker"]): 
            CIK = str(ticker_cik_map[str(i)]["cik_str"])
            return CIK
    
    # return error code if ticker not found
    return constants.ERRORCODE

# gets company facts data for given stock ticker using SEC API
# @param ticker - stock ticker
# @return dictionary - JSON as dictionary containing company facts data
def get_company_facts_json_from_ticker(ticker):
    cik = get_cik_from_ticker(ticker)

    # check if CIK is valid
    assert(cik != constants.ERRORCODE), "Unable to find company CIK from ticker"
    assert(cik.isdigit and len(cik) >= 5 and len(cik) <= 7), "Invalid CIK number"
    
    # set CIK Prefix (note: can be either 3 or 4 zeros) 
    cik_prefix = "CIK000"
    while((len(cik)+len(cik_prefix)) < 13): 
        cik_prefix += "0"

    # create company CIK and build URL
    company_cik = cik_prefix + str(cik)
    company_facts_url = constants.Company_Facts_link.replace("$$CIK$$", company_cik)

    # get company facts from SEC Edgar API
    company_facts_data = get_json_from_url_link(company_facts_url)
    assert(company_facts_data != constants.ERRORCODE), "Unable to fetch Company Facts JSON"
    print("Loaded SEC JSON data sucessfully!")
    return company_facts_data

if __name__ == "__main__":
    get_company_facts_json_from_ticker("aapl")
    get_stock_price_from_ticker("aapl")
