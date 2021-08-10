from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
import constants
from util import percent_to_dec
from util import get_excel_cell

class spreadsheet:
    "Class to create excel sheet containing DCF model"
    def __init__(self, ticker, inputs, num_years=10, rates=constants.Default_rates):
        self.ticker = ticker
        self.num_years = num_years
        self.rates = rates
        self.inputs = inputs

    # sets revenue and revenue growth rate rows for the given worksheet
    # @param ws - reference to excel worksheet
    def set_revenue(self, ws):
        row = constants.excel_list_row_numbers["Revenue"]
        revenue_cell = get_excel_cell(2, row)
        growth_cell = get_excel_cell(2, row+1)
        
        ws[revenue_cell] = int(self.inputs["Revenue"][0])
        ws[growth_cell] = 0
        
        for i in range(1, self.num_years+2):  
            growth_cell = get_excel_cell(2+i, row+1)
            revenue_cell = get_excel_cell(2+i, row)
            prev_revenue_cell = get_excel_cell(1+i, row)

            if(i == 1): 
                ws[growth_cell] = percent_to_dec(self.rates["growth_rate_1_year"])
            elif(i >= 2 and i <= 5):
                ws[growth_cell] = percent_to_dec(self.rates["growth_rate_2_to_5_year"])
            elif(i >= self.num_years+1): 
                ws[growth_cell] = percent_to_dec(self.rates['risk_free_rate'])
            else: 
                fraction = percent_to_dec(self.rates["growth_rate_2_to_5_year"] - self.rates["risk_free_rate"])
                growth_rate = fraction*(self.num_years-i)/(self.num_years - 5)
                ws[growth_cell] = growth_rate + percent_to_dec(self.rates["risk_free_rate"])
            ws[revenue_cell] = "=(1+" + growth_cell + ")*" + prev_revenue_cell

    # sets operating income, costs and expenses and operating
    # margin rows for the given worksheet
    # @param ws - reference to excel worksheet    
    def set_operating_income(self, ws): 
        row = constants.excel_list_row_numbers["Costs and Expenses"]
        revenue_row = constants.excel_list_row_numbers["Revenue"]
        
        costs_and_expenses_cell = get_excel_cell(2, row)
        operating_income_cell = get_excel_cell(2, row+1)
        operating_margin_cell = get_excel_cell(2, row+2)
        revenue_cell = get_excel_cell(2, revenue_row)

        c_and_e_theoretical = int(self.inputs["Revenue"][0]) - int(self.inputs["Operating Income"][0])
        if(c_and_e_theoretical != int(self.inputs['COGS'][0])):
            self.inputs["COGS"][0] = c_and_e_theoretical

        ws[costs_and_expenses_cell] = int(self.inputs['COGS'][0])
        ws[operating_income_cell] = int(self.inputs['Operating Income'][0])
        ws[operating_margin_cell] = ws[operating_income_cell].value/ws[revenue_cell].value

        for i in range(1, self.num_years+2):
            costs_and_expenses_cell = get_excel_cell(2+i, row)
            operating_income_cell = get_excel_cell(2+i, row+1)
            operating_margin_cell = get_excel_cell(2+i, row+2)
            revenue_cell = get_excel_cell(2+i, revenue_row)            
            
            if(i == 1): 
                ws[operating_margin_cell] = percent_to_dec(self.rates["operating_margin_1_year"])
            elif(i == self.num_years+1): 
                ws[operating_margin_cell] = percent_to_dec(self.rates["target_pretax_operating_margin"])
            elif(i > 1 and i <= 5): 
                target = self.rates["target_pretax_operating_margin"]
                year_1 = self.rates["operating_margin_1_year"]
                margin =  target - ((target-year_1)/self.num_years)*(self.num_years-i)
                ws[operating_margin_cell] = percent_to_dec(margin)
            else: 
                target = self.rates["target_pretax_operating_margin"]
                initial = ws[get_excel_cell(2, row+2)].value
                margin =  target - ((target-initial)/self.num_years)*(self.num_years-i)
                ws[operating_margin_cell] = percent_to_dec(margin)
            ws[operating_income_cell] = "=" + revenue_cell + "*" + operating_margin_cell
            ws[costs_and_expenses_cell] = "=" + revenue_cell + "-" + operating_income_cell
    
    # sets tax and tax rate rows for the given worksheet
    # @param ws - reference to excel worksheet
    def set_taxes(self, ws): 
        row = constants.excel_list_row_numbers["Taxes"]
        oi_row = constants.excel_list_row_numbers["Operating Income"]
        
        taxes_cell = get_excel_cell(2, row)
        initial_trc = get_excel_cell(2, row+1)
        final_trc = get_excel_cell(2+self.num_years+1, row+1)
        operating_income_cell = get_excel_cell(2, oi_row)

        ws[taxes_cell] = self.inputs["Taxes"][0]
        initial_tax_rate = ws[taxes_cell].value/ws[operating_income_cell].value
        marginal_tax_rate = percent_to_dec(self.rates["marginal_tax_rate"])
        ws[final_trc] = marginal_tax_rate
        
        if(initial_tax_rate < 0):
            ws[initial_trc] = 0
        elif(initial_tax_rate > 1): 
            ws[initial_trc] = marginal_tax_rate
        else: 
            ws[initial_trc] = initial_tax_rate
        
        for i in range(1, self.num_years+2): 
            tax_rate_cell = get_excel_cell(2+i, row+1)
            prev_tax_rate_cell = get_excel_cell(2+i-1, row+1)
            
            taxes_cell = get_excel_cell(2+i, row)
            operating_income_cell = get_excel_cell(2+i, oi_row)
            
            if(i <= self.num_years/2):
                ws[tax_rate_cell] = "=" + prev_tax_rate_cell 
            elif(i < self.num_years+1): 
                diff = "ABS(" + initial_trc + "-" + final_trc + ")/" + str(self.num_years/2)
                ws[tax_rate_cell] = "=" + prev_tax_rate_cell + "+" + diff
            ws[taxes_cell] = "=" + tax_rate_cell + "*" + operating_income_cell
    
    # sets net income row for the given worksheet
    # @param ws - reference to excel worksheet
    def set_net_income(self, ws): 
        row = constants.excel_list_row_numbers["Net Income"]
        oi_row = constants.excel_list_row_numbers["Operating Income"]
        tax_rate_row = constants.excel_list_row_numbers["Tax Rate"]

        for i in range(self.num_years+2): 
            net_income_cell = get_excel_cell(2+i, row)
            tax_rate_cell = get_excel_cell(2+i, tax_rate_row)
            operating_income_cell = get_excel_cell(2+i, oi_row)
            if(i == 0): 
                ws[net_income_cell] = self.inputs["Net Income"][0]
            else:
                ws[net_income_cell] = "=" + operating_income_cell + "*(1-" + tax_rate_cell + ")"
    
    # sets sales to capital ratio row for the given worksheet
    # @param ws - reference to excel worksheet
    def set_s_to_c_ratio(self, ws): 
        row = constants.excel_list_row_numbers["Sales to Capital Ratio"]
        for i in range(1, self.num_years+2):
            cell = get_excel_cell(2+i, row)
            if(i == 1): 
                ws[cell] = self.rates["sales_to_capital_ratio_1_year"]
            elif(i > 1 and i < 6): 
                ws[cell] = self.rates["sales_to_capital_ratio_2_to_5_year"]
            else: 
                ws[cell] = self.rates["sales_to_capital_ratio_6_to_10_year"]
    
    # sets sales to capital ratio row for the given worksheet
    # @param ws - reference to excel worksheet
    def set_reinvestment(self, ws): 
        row = constants.excel_list_row_numbers["Reinvestment"]
        s_to_c_row = constants.excel_list_row_numbers["Sales to Capital Ratio"]
        rev_row = constants.excel_list_row_numbers["Revenue"]

        for i in range(1, self.num_years+2): 
            rev_cell_1 = get_excel_cell(1+i, rev_row)
            rev_cell_2 = get_excel_cell(2+i, rev_row)
            s_to_c_cell = get_excel_cell(2+i, s_to_c_row)
            reinvestment_cell = get_excel_cell(2+i, row)

            formula = "(" + rev_cell_2 + "-" + rev_cell_1 + ")/" + s_to_c_cell     
            if(i == 1): 
                conditional = "=if(" + rev_cell_2 + ">" + rev_cell_1 + ","
                ws[reinvestment_cell] = conditional + formula + ",0)"
            else: 
                ws[reinvestment_cell] = "=" + formula

    # sets free cash flow rows for the given worksheet
    # @param ws - reference to excel worksheet
    def set_free_cash_flow(self, ws): 
        row = constants.excel_list_row_numbers["FCFF"]
        net_income_row = constants.excel_list_row_numbers["Net Income"]
        reinvestment_row = constants.excel_list_row_numbers["Reinvestment"]

        for i in range(1, self.num_years+2): 
            fcf_cell = get_excel_cell(2+i, row)
            net_income_cell = get_excel_cell(2+i, net_income_row)
            reinvestment_cell = get_excel_cell(2+i, reinvestment_row)
            ws[fcf_cell] = "=" + net_income_cell + "-" + reinvestment_cell
    
    # sets cost of capital row for the given worksheet
    # @param ws - reference to excel worksheet
    def set_cost_of_capital(self, ws):
        row = constants.excel_list_row_numbers["Cost of Capital"]
        for i in range(1, self.num_years+2): 
            cell = get_excel_cell(2+i, row)
            ws[cell] = percent_to_dec(self.rates["initial_cost_of_capital"])
    
    # sets discount rate row for the given worksheet
    # @param ws - reference to excel worksheet
    def set_cumulative_discount_factor(self, ws): 
        row = constants.excel_list_row_numbers["Cumulative Discount Factor"]
        coc_row = constants.excel_list_row_numbers["Cost of Capital"]
        for i in range(0, self.num_years+2): 
            df_cell = get_excel_cell(2+i, row)
            if(i == 0): 
                ws[df_cell] = 1. 
            else: 
                prev_df_cell = get_excel_cell(1+i, row)
                coc_cell = get_excel_cell(2+i, coc_row)
                ws[df_cell] = "=" + prev_df_cell + "/(1+" + coc_cell + ")"
    
    # sets present value row for the given worksheet
    # @param ws - reference to excel worksheet
    def set_present_value(self, ws): 
        row = constants.excel_list_row_numbers["PV (FCFF)"]
        fcf_row = constants.excel_list_row_numbers["FCFF"]
        df_row = constants.excel_list_row_numbers["Cumulative Discount Factor"]
        
        for i in range(1, self.num_years+2): 
            cell = get_excel_cell(2+i, row)
            fcf_cell = get_excel_cell(2+i, fcf_row)
            df_cell = get_excel_cell(2+i, df_row)
            ws[cell] = "=" + fcf_cell + "*" + df_cell

    # sets npv outputs cells (terminal value, cash flow etc.) for the given worksheet
    # @param ws - reference to excel worksheet
    def set_npv_outputs(self, ws): 
        last_col = self.num_years+3
        fcf_row = constants.excel_list_row_numbers["FCFF"]
        coc_row = constants.excel_list_row_numbers["Cost of Capital"]
        rev_growth_row = constants.excel_list_row_numbers["Revenue Growth Rate"]
        cdf_row = constants.excel_list_row_numbers["Cumulative Discount Factor"]
        pv_row = constants.excel_list_row_numbers["PV (FCFF)"]
        
        tcf_row = constants.excel_singular_row_numbers["Terminal Cash Flow"]
        tcoc_row = constants.excel_singular_row_numbers["Terminal Cost of Capital"]
        tv_row = constants.excel_singular_row_numbers["Terminal Value"]
        pv_terminal_row = constants.excel_singular_row_numbers["PV (Terminal Value)"]
        pv_cashflows_row = constants.excel_singular_row_numbers["PV (Cashflows)"]
        pv_sum_row = constants.excel_singular_row_numbers["Sum of PV"]
        
        ws[get_excel_cell(2, tcf_row)] = "=" + get_excel_cell(last_col, fcf_row)
        ws[get_excel_cell(2, tcoc_row)] = "=" + get_excel_cell(last_col, coc_row)
        ws[get_excel_cell(2, tv_row)] = "=" + get_excel_cell(2, tcf_row) + "/(" + \
            get_excel_cell(2, tcoc_row) + "-" + get_excel_cell(last_col, rev_growth_row) + ")"
        ws[get_excel_cell(2, pv_terminal_row)] = "=" + get_excel_cell(2, tv_row) + "*" \
            + get_excel_cell(last_col, cdf_row)
        ws[get_excel_cell(2, pv_cashflows_row)] = "=SUM(" + get_excel_cell(2, pv_row) \
            + ":" + get_excel_cell(last_col, pv_row) + ")" 
        ws[get_excel_cell(2, pv_sum_row)] = "=" + get_excel_cell(2, pv_terminal_row) \
            + "+" + get_excel_cell(2, pv_cashflows_row)
    
    # sets overall npv valuation for the given worksheet
    # @param ws - reference to excel worksheet
    def set_npv_valutaion(self, ws): 
        cash_row = constants.excel_singular_row_numbers["Cash"]
        debt_row = constants.excel_singular_row_numbers["Debt"]
        pv_sum_row = constants.excel_singular_row_numbers["Sum of PV"]
        voe_row = constants.excel_singular_row_numbers["Value of Equity"]
        shares_row = constants.excel_singular_row_numbers["Common Shares Outstanding"]
        vps_row = constants.excel_singular_row_numbers["Value per Share"]
        price_row = constants.excel_singular_row_numbers["Current Price"]
        
        ws[get_excel_cell(2, cash_row)] = self.inputs["Cash"][0] + \
            self.inputs["Current Marketable Securities"][0] + \
            self.inputs["Noncurrent Marketable Securities"][0]
        ws[get_excel_cell(2, debt_row)] = self.inputs["Debt"][0]
        ws[get_excel_cell(2, voe_row)] = "=" + get_excel_cell(2, pv_sum_row) + \
            "+" + get_excel_cell(2, cash_row) + "-" + get_excel_cell(2, debt_row)
        ws[get_excel_cell(2, shares_row)] = self.inputs["Shares"][0]
        ws[get_excel_cell(2, vps_row)] = "=" + get_excel_cell(2, voe_row) + \
            "/" + get_excel_cell(2, shares_row)
        ws[get_excel_cell(2, price_row)] = self.inputs["Price"][0]

    # sets row and column names for the given worksheet
    # @param ws - reference to excel worksheet
    def set_row_and_col_names(self, ws):
        row_names = {**constants.excel_list_row_numbers, **constants.excel_singular_row_numbers}
        for k, v in row_names.items(): 
            ws[get_excel_cell(1, v)] = k
        
        for i in range(self.num_years+2): 
            if(i == 0): 
                ws[get_excel_cell(i+2, 1)] = "Base Year"
            elif(i == self.num_years+1): 
                ws[get_excel_cell(i+2, 1)] = "Terminal Year"
            else: 
                ws[get_excel_cell(i+2, 1)] = "Year " + str(i)

    # formats excel sheet worksheet
    # @param ws - reference to excel worksheet
    def format_excel_sheet(self, ws): 
        col = get_column_letter(1)
        ws.column_dimensions[get_column_letter(1)].width = constants.excel_column_widths["big"]
        for i in range(1, self.num_years+3):
            col = get_column_letter(i+1)
            ws.column_dimensions[col].width = constants.excel_column_widths["small"]

        percentages = ['Revenue Growth Rate', 'Operating Margin', 'Tax Rate', 
            'Cost of Capital', 'Cumulative Discount Factor', 'Terminal Cost of Capital']
        small_numbers = ['Value per Share', 'Current Price']
        
        for k, v in constants.excel_list_row_numbers.items():
            for j in range(2, self.num_years+4):
                if(k in percentages):   
                    ws[get_excel_cell(j, v)].number_format = '0.00%'
                else: 
                    ws[get_excel_cell(j, v)].number_format = '#,##0_);(#,##0)'
        
        for k, v in constants.excel_singular_row_numbers.items():
            if(k in small_numbers): 
                ws[get_excel_cell(2, v)].number_format = '0.00'
            elif(k in percentages): 
                ws[get_excel_cell(2, v)].number_format = '0.00%'
            else: 
                ws[get_excel_cell(2, v)].number_format = '#,##0_);(#,##0)'

    # sets all cashflows for the DCF in the given worksheet 
    # @param ws - reference to excel worksheet
    def set_cashflows(self, ws): 
        self.set_revenue(ws)
        self.set_operating_income(ws)
        self.set_taxes(ws)
        self.set_net_income(ws)
        self.set_s_to_c_ratio(ws)
        self.set_reinvestment(ws)
        self.set_free_cash_flow(ws)
    
    # sets all npv output cells for the given worksheet
    # @param ws - reference to excel worksheet
    def set_npv(self, ws):
        self.set_cost_of_capital(ws)
        self.set_cumulative_discount_factor(ws)
        self.set_present_value(ws)
        self.set_npv_outputs(ws)
        self.set_npv_valutaion(ws)

    # creates DCF model excel sheet in the given file
    # @param filename - path to new excel file
    def create_excel_sheet(self, filename):
        # add .xlsx extension if it doesn't exist already
        if((not filename.endswith(".xlsx")) and (not filename.endswith(".xls"))): 
            print("Adding .xlsx extension to given filename")
            filename += ".xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Valuation Output - " + self.ticker.upper()
        
        self.set_cashflows(ws)
        self.set_npv(ws)
        self.set_row_and_col_names(ws)
        self.format_excel_sheet(ws)

        wb.save(filename)
        print('DCF Spreadsheet generated succesfully!')
        print('Filename:', filename)
